import io
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

def fetch_meta(raw_url: str) -> dict:
    url = raw_url if raw_url.startswith(("http://", "https://")) else "https://" + raw_url
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")

        title_tag = soup.find("title")
        title = title_tag.get_text(strip=True) if title_tag else ""

        desc_tag = (
            soup.find("meta", attrs={"name": "description"}) or
            soup.find("meta", attrs={"name": "Description"}) or
            soup.find("meta", attrs={"property": "og:description"})
        )
        description = desc_tag.get("content", "").strip() if desc_tag else ""

        return {"url": url, "meta_title": title, "meta_description": description, "error": ""}

    except Exception as e:
        return {"url": url, "meta_title": "", "meta_description": "", "error": str(e)}

def seo_status(length: int, lo: int, hi: int) -> str:
    if length == 0:   return "Missing"
    if length < lo:   return "Too Short"
    if length > hi:   return "Too Long"
    return "OK"

def build_excel(results: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meta Scraper Results"

    header_fill = PatternFill("solid", fgColor="7C6AFF")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="2A2A3A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["URL", "Meta Title", "Title Length", "Title Status",
               "Meta Description", "Desc Length", "Desc Status", "Error"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 22

    fill_ok = PatternFill("solid", fgColor="D1FAE5")
    fill_warn = PatternFill("solid", fgColor="FEF9C3")
    fill_bad = PatternFill("solid", fgColor="FEE2E2")

    def status_fill(status: str):
        if status == "OK": return fill_ok
        if status == "Missing": return fill_bad
        return fill_warn

    for r in results:
        tl = len(r.get("meta_title", ""))
        dl = len(r.get("meta_description", ""))
        ts = seo_status(tl, 30, 60)
        ds = seo_status(dl, 70, 160)
        row = [r.get("url", ""), r.get("meta_title", ""), tl, ts,
               r.get("meta_description", ""), dl, ds, r.get("error", "")]
        ws.append(row)
        row_idx = ws.max_row

        ws.cell(row=row_idx, column=4).fill = status_fill(ts)
        ws.cell(row=row_idx, column=7).fill = status_fill(ds)
        if r.get("error"):
            ws.cell(row=row_idx, column=8).font = Font(color="DC2626", bold=True)

        for col_idx in [1, 2, 5]:
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    col_max_widths = [60, 40, 12, 14, 60, 12, 14, 40]
    for idx, max_w in enumerate(col_max_widths, start=1):
        col_letter = get_column_letter(idx)
        max_seen = max(
            (len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_seen + 4, max_w)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

@app.route('/api/scrape', methods=['POST'])
def scrape():
    data = request.json
    urls = data.get('urls', [])
    
    if not urls:
        return jsonify({"error": "No URLs provided"}), 400

    results = []
    for url in urls:
        if url.strip():
            results.append(fetch_meta(url.strip()))
            time.sleep(0.3)

    return jsonify({"results": results})

@app.route('/api/export', methods=['POST'])
def export():
    data = request.json
    results = data.get('results', [])
    
    if not results:
        return jsonify({"error": "No results provided"}), 400

    excel_data = build_excel(results)
    
    return send_file(
        io.BytesIO(excel_data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='meta_data.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
