"""
Meta Tag Scraper — Python/Streamlit version
Converted from React JSX (meta-scraper-ui.jsx)

Run with:
    streamlit run meta_scraper.py
"""

import io
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import requests
import streamlit as st
from bs4 import BeautifulSoup

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Meta Scraper",
    page_icon="⚡",
    layout="centered",
)

# ── Custom CSS (mirrors the dark theme from the JSX) ─────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: #0a0a0f;
    color: #e8e8f0;
}

/* Hide default streamlit branding */
#MainMenu, footer, header { visibility: hidden; }

/* App background */
.stApp {
    background: #0a0a0f;
    background-image:
        radial-gradient(ellipse 60% 40% at 70% 10%, rgba(124,106,255,0.08) 0%, transparent 60%),
        radial-gradient(ellipse 40% 30% at 10% 80%, rgba(255,106,142,0.06) 0%, transparent 50%);
}

/* Inputs */
.stTextArea textarea {
    background: #111118 !important;
    border: 1px solid #2a2a3a !important;
    color: #e8e8f0 !important;
    font-family: 'DM Mono', monospace !important;
    font-size: 13px !important;
    border-radius: 10px !important;
}
.stTextArea textarea:focus {
    border-color: #7c6aff !important;
    box-shadow: 0 0 0 1px #7c6aff !important;
}

/* Primary button */
.stButton > button[kind="primary"] {
    background: #7c6aff !important;
    border: none !important;
    color: white !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 10px 24px !important;
    transition: all 0.2s !important;
}
.stButton > button[kind="primary"]:hover {
    background: #9580ff !important;
    box-shadow: 0 8px 20px rgba(124,106,255,0.3) !important;
    transform: translateY(-1px) !important;
}

/* Secondary button */
.stButton > button[kind="secondary"] {
    background: transparent !important;
    border: 1px solid #2a2a3a !important;
    color: #6b6b85 !important;
    font-family: 'Syne', sans-serif !important;
    border-radius: 10px !important;
}
.stButton > button[kind="secondary"]:hover {
    border-color: #f87171 !important;
    color: #f87171 !important;
}

/* Download button */
.stDownloadButton > button {
    background: rgba(74,222,128,0.1) !important;
    border: 1px solid rgba(74,222,128,0.25) !important;
    color: #4ade80 !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
}
.stDownloadButton > button:hover {
    background: rgba(74,222,128,0.2) !important;
}

/* Result cards */
.result-card {
    background: #16161f;
    border: 1px solid #2a2a3a;
    border-radius: 14px;
    padding: 18px 20px;
    margin-bottom: 12px;
    border-left: 3px solid #7c6aff;
}
.result-card.has-error { border-left-color: #f87171; }
.result-url {
    font-family: 'DM Mono', monospace;
    font-size: 12px;
    color: #7c6aff;
    margin-bottom: 14px;
    word-break: break-all;
}
.meta-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #6b6b85;
    margin-bottom: 4px;
}
.meta-value {
    font-size: 14px;
    color: #e8e8f0;
    background: #111118;
    border-radius: 8px;
    padding: 8px 12px;
    margin-bottom: 12px;
}
.meta-value.empty { color: #6b6b85; font-style: italic; }
.meta-value.err   { color: #f87171; font-family: 'DM Mono', monospace; font-size:12px; }

.pill {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 20px;
    font-family: 'DM Mono', monospace;
    font-size: 11px;
    font-weight: 500;
    margin-left: 8px;
}
.pill-ok   { background: rgba(74,222,128,0.12);  color: #4ade80;  border: 1px solid rgba(74,222,128,0.25); }
.pill-warn { background: rgba(251,191,36,0.12);  color: #fbbf24;  border: 1px solid rgba(251,191,36,0.25); }
.pill-bad  { background: rgba(248,113,113,0.12); color: #f87171;  border: 1px solid rgba(248,113,113,0.25); }

.dot { display: inline-block; width: 8px; height: 8px; border-radius: 50%; margin-right: 8px; }
.dot-ok  { background: #4ade80; }
.dot-err { background: #f87171; }
</style>
""", unsafe_allow_html=True)


# ── SEO helpers ───────────────────────────────────────────────────────────────
def seo_pill(length: int, lo: int, hi: int) -> str:
    """Return an HTML pill badge based on character length."""
    if length == 0:
        return '<span class="pill pill-bad">Missing</span>'
    if length < lo:
        return f'<span class="pill pill-warn">{length} · Too Short</span>'
    if length > hi:
        return f'<span class="pill pill-warn">{length} · Too Long</span>'
    return f'<span class="pill pill-ok">{length} · OK</span>'


def seo_status(length: int, lo: int, hi: int) -> str:
    if length == 0:   return "Missing"
    if length < lo:   return "Too Short"
    if length > hi:   return "Too Long"
    return "OK"


# ── Core scraper (replaces fetchMeta in JSX) ──────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

def fetch_meta(raw_url: str) -> dict:
    """
    Fetch a URL and extract <title> + meta description.
    Returns a dict with: url, meta_title, meta_description, error
    """
    url = raw_url if raw_url.startswith(("http://", "https://")) else "https://" + raw_url
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")

        # Title
        title_tag = soup.find("title")
        title = title_tag.get_text(strip=True) if title_tag else ""

        # Meta description (same priority as JSX)
        desc_tag = (
            soup.find("meta", attrs={"name": "description"}) or
            soup.find("meta", attrs={"name": "Description"}) or
            soup.find("meta", attrs={"property": "og:description"})
        )
        description = desc_tag.get("content", "").strip() if desc_tag else ""

        return {"url": url, "meta_title": title, "meta_description": description, "error": ""}

    except Exception as e:
        return {"url": url, "meta_title": "", "meta_description": "", "error": str(e)}


# ── Excel export helper ───────────────────────────────────────────────────────
def build_excel(results: list) -> bytes:
    """Convert results list to a styled Excel .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meta Scraper Results"

    # ── Header style
    header_fill = PatternFill("solid", fgColor="7C6AFF")   # accent purple
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="2A2A3A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["URL", "Meta Title", "Title Length", "Title Status",
               "Meta Description", "Desc Length", "Desc Status", "Error"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 22

    # ── Status fill colours
    fill_ok   = PatternFill("solid", fgColor="D1FAE5")   # light green
    fill_warn = PatternFill("solid", fgColor="FEF9C3")   # light yellow
    fill_bad  = PatternFill("solid", fgColor="FEE2E2")   # light red

    def status_fill(status: str):
        if status == "OK":      return fill_ok
        if status == "Missing": return fill_bad
        return fill_warn

    # ── Data rows
    for r in results:
        tl = len(r["meta_title"])
        dl = len(r["meta_description"])
        ts = seo_status(tl, 30, 60)
        ds = seo_status(dl, 70, 160)
        row = [r["url"], r["meta_title"], tl, ts,
               r["meta_description"], dl, ds, r.get("error", "")]
        ws.append(row)
        row_idx = ws.max_row

        # Colour the status cells
        ws.cell(row=row_idx, column=4).fill = status_fill(ts)   # Title Status
        ws.cell(row=row_idx, column=7).fill = status_fill(ds)   # Desc Status
        if r.get("error"):
            ws.cell(row=row_idx, column=8).font = Font(color="DC2626", bold=True)

        # Wrap long text
        for col_idx in [1, 2, 5]:
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Auto column widths
    col_max_widths = [60, 40, 12, 14, 60, 12, 14, 40]
    for idx, max_w in enumerate(col_max_widths, start=1):
        col_letter = get_column_letter(idx)
        # sample the data to estimate width
        max_seen = max(
            (len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_seen + 4, max_w)

    # ── Freeze top row
    ws.freeze_panes = "A2"

    # ── Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Session state init ────────────────────────────────────────────────────────
if "urls" not in st.session_state:
    st.session_state.urls = []
if "results" not in st.session_state:
    st.session_state.results = []


# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="margin-bottom:40px">
  <div style="display:inline-flex;align-items:center;gap:6px;
              background:rgba(124,106,255,0.12);border:1px solid rgba(124,106,255,0.25);
              border-radius:20px;padding:4px 12px;font-family:'DM Mono',monospace;
              font-size:11px;color:#7c6aff;letter-spacing:0.05em;margin-bottom:16px">
    ● Meta Scraper
  </div>
  <h1 style="font-size:clamp(28px,5vw,42px);font-weight:800;line-height:1.1;
             letter-spacing:-0.03em;margin:0">
    Extract <span style="color:#7c6aff">Meta Tags</span><br/>from any URL
  </h1>
  <p style="color:#6b6b85;font-size:14px;margin-top:8px;font-family:'DM Mono',monospace;font-weight:300">
    // paste URLs → scan → export CSV
  </p>
</div>
""", unsafe_allow_html=True)


# ── Input card ────────────────────────────────────────────────────────────────
with st.container():
    st.markdown('<div class="meta-label">URLs to scan</div>', unsafe_allow_html=True)
    raw_input = st.text_area(
        label="urls",
        label_visibility="collapsed",
        placeholder="https://example.com\nhttps://another.com",
        height=100,
        key="url_input",
    )

    col1, col2, col3 = st.columns([2, 1, 3])
    with col1:
        if st.button("⚡ Scan All", type="primary", use_container_width=True):
            # Parse URLs from textarea
            new_urls = [u.strip() for u in raw_input.replace(",", "\n").splitlines() if u.strip()]
            all_urls = list(dict.fromkeys(st.session_state.urls + new_urls))  # dedupe
            st.session_state.urls = all_urls

            if all_urls:
                st.session_state.results = []
                progress_bar = st.progress(0, text="Starting…")
                results_out = []

                for i, url in enumerate(all_urls):
                    progress_bar.progress(
                        (i + 1) / len(all_urls),
                        text=f"Scanning {i+1}/{len(all_urls)} — {url}"
                    )
                    results_out.append(fetch_meta(url))
                    time.sleep(0.3)   # polite delay

                progress_bar.empty()
                st.session_state.results = results_out

    with col2:
        if st.button("Clear", type="secondary", use_container_width=True):
            st.session_state.urls = []
            st.session_state.results = []
            st.rerun()

    # URL count badge
    if st.session_state.urls:
        n = len(st.session_state.urls)
        st.markdown(
            f'<p style="font-family:\'DM Mono\',monospace;font-size:12px;color:#6b6b85">'
            f'{n} URL{"s" if n != 1 else ""} queued</p>',
            unsafe_allow_html=True
        )


# ── Results ───────────────────────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results

    col_title, col_export = st.columns([3, 1])
    with col_title:
        n = len(results)
        st.markdown(
            f'<div style="font-size:13px;font-weight:700;letter-spacing:0.05em;'
            f'text-transform:uppercase;color:#6b6b85;margin:24px 0 16px">'
            f'Results — {n} page{"s" if n != 1 else ""}</div>',
            unsafe_allow_html=True
        )
    with col_export:
        st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)
        st.download_button(
            label="↓ Export Excel (.xlsx)",
            data=build_excel(results),
            file_name="meta_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    for r in results:
        tl = len(r["meta_title"])
        dl = len(r["meta_description"])
        has_error = bool(r["error"])
        card_class = "result-card has-error" if has_error else "result-card"
        dot_class  = "dot dot-err" if has_error else "dot dot-ok"

        if has_error:
            st.markdown(f"""
            <div class="{card_class}">
              <div class="result-url"><span class="{dot_class}"></span>{r['url']}</div>
              <div class="meta-label">Error</div>
              <div class="meta-value err">{r['error']}</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            title_display = r["meta_title"] or "No title tag found"
            desc_display  = r["meta_description"] or "No meta description found"
            title_empty   = "" if r["meta_title"] else " empty"
            desc_empty    = "" if r["meta_description"] else " empty"

            st.markdown(f"""
            <div class="{card_class}">
              <div class="result-url"><span class="{dot_class}"></span>{r['url']}</div>

              <div class="meta-label">Meta Title {seo_pill(tl, 30, 60)}</div>
              <div class="meta-value{title_empty}">{title_display}</div>

              <div class="meta-label">Meta Description {seo_pill(dl, 70, 160)}</div>
              <div class="meta-value{desc_empty}">{desc_display}</div>
            </div>
            """, unsafe_allow_html=True)
